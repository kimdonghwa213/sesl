from __future__ import print_function
import openpyxl
import math
import pandas as pd
import numpy as np
from matplotlib import pyplot as plt
from CoolProp import AbstractState
from CoolProp.CoolProp import PhaseSI, PropsSI, get_global_param_string
import CoolProp.CoolProp as CoolProp

# 기본 상수 정의
P_atm = 101325  # [K]
T_mean = 300  # [K]
rhoAir = PropsSI("D", "T", T_mean, "P", P_atm, "air")  # 1.1769955883877592 [kg/m3]
rhoWater = PropsSI("D", "T", T_mean, "P", P_atm, "water")  # 996.5569352652021 [kg/m3]
cAir = PropsSI("C", "T", T_mean, "P", P_atm, "air") / 1000  # 1.0063739076641027 [kJ/kgK]
cWater = PropsSI("C", "T", T_mean, "P", P_atm, "water") / 1000  # 4.1806357765560715 [kJ/kgK]
pi = 3.14

# 먼저 시간 배열 정의
hour = np.ones(8760)

# 그 다음 나머지 배열들 초기화
Abs = np.ones(1470)
wavelength = np.linspace(0.3, 15, 1470)
Toutdoor = 300 * np.ones(len(hour))  # [K]
Troom = 300 * np.ones(len(hour))
Vwind = np.ones(8760)

hconv = np.ones(8760)
qTotalHouse = np.zeros(len(hour))  # [W]
qInHouse = np.zeros(len(hour))  # [W]
radSolar = np.zeros(len(hour))
qRad = np.zeros(len(hour))
qRoof = np.zeros(len(hour))  # [W]
qFloor = np.zeros(len(hour))  # [W]
qVent = np.zeros(len(hour))  # [W]
qSideWall = np.zeros(len(hour))  # [W]
qFrontBack = np.zeros(len(hour))
qWindow = np.zeros(len(hour))  # 창문 열손실 [W]
qHeating = np.zeros(len(hour))  # [W]
qCooling = np.zeros(len(hour))  # [W]
PW = np.zeros(len(hour))  # [W]
mode = np.zeros(len(hour))  # [W]

# 엑셀 파일 열기 (openpyxl 사용)
wb = openpyxl.load_workbook('TMY3_Gangnung.xlsx')
sheet = wb.active

# 데이터를 엑셀 파일에서 읽어오기
for i in range(0, len(hour)):
    Toutdoor[i] = sheet.cell(row=i + 1, column=2).value  # 배열은 0부터, 엑셀은 1부터 시작
    Vwind[i] = sheet.cell(row=i + 1, column=3).value
    PW[i] = sheet.cell(row=i + 1, column=4).value
    radSolar[i] = sheet.cell(row=i + 1, column=5).value

# 온실 크기 및 기타 변수 설정
lengthHouse = 70  # [m]
widthHouse = 8.5  # [m]
heightHouse = 4.0  # [m]
fracSolarWindow = 0.5
fracLossWindow = 1
nFloor = 1
transGlass = 0.85
areaRoof = 696.36
areaFrontBack = 64.5
areaSideWall = 604.8
rGround = 3.0
areaHouse = lengthHouse * widthHouse
surfaceHouse = (widthHouse * heightHouse + lengthHouse * heightHouse) * 2
volumeHouse = areaHouse * heightHouse * nFloor
mHouse = areaHouse * heightHouse * nFloor * rhoAir
ht = 5.7 # 열 관류율 [㎉/㎡·h·℃]
hs = 0.244 # 지표면 전열 계수 [㎉/㎡·h·℃]
fr = 0.3 # 보온 피복재의 열 절감율
hv = 0.2
Agh = 1365.66

rRoof = 0.3
rFloor = 2
rSideWall = 0.3

# 환기 관련 변수 설정
ACH = 1  # Air Changes per Hour (시간당 환기횟수)
ventilation_rate = (ACH * volumeHouse) / 3600  # [m³/s]
m_vent = ventilation_rate * rhoAir  # 환기 공기 질량유량 [kg/s]

# 실내 온도 초기화
Troom_initial = 13 + 273
Troom[0] = Troom_initial
Tground = 15 + 273
Tset_heating = 20 + 273  # 난방 설정 온도 (16°C)
Tset_cooling = 30 + 273  # 냉방 설정 온도 (28°C)

# Tsolair2 계산을 위한 변수 정의
alpha_roof = 0.15  # 지붕 표면 일사흡수율 [-]

def calculate_Tsolair2(k, Toutdoor, radSolar):
    Tsolair2 = Toutdoor[k] + ((alpha_roof * radSolar[k]) / 17)  # 등가 외기온도 [K]
    return Tsolair2

# Tsolair2 계산
Tsolair2 = np.zeros(len(hour))
for k in range(len(hour)):
    Tsolair2[k] = calculate_Tsolair2(k, Toutdoor, radSolar)

# 열 부하 계산
for k in range(len(hour)):
    # 1. 먼저 현재 시간의 실내온도 계산 (k > 0일 때)
    if k > 0:
        dT = (qTotalHouse[k-1] ) / (mHouse * cAir )  # 온도 변화
        Troom[k] = Troom[k-1] + dT  # 새로운 실내온도

    # 2. 계산된 실내온도로 열부하 계산
    # qVent[k] = (Tsolair2[k] - Troom[k]) * hv * Agh * 0.001163
    # qRad[k] = fracSolarWindow * transGlass * areaHouse * radSolar[k] / 1000
    # qRoof[k] = (Tsolair2[k] - Troom[k]) * ht * areaRoof * (1 - fr) * 0.001163   #kWh
    # qFloor[k] = (Tground - Troom[k]) * hs * areaHouse * 0.001163
    # qSideWall[k] = (Tsolair2[k] - Troom[k]) * ht * areaSideWall * (1 - fr) * 0.001163
    # qFrontBack[k] = (Tsolair2[k] - Troom[k]) * ht * areaFrontBack * (1 - fr) * 0.001163

    qRad[k] = (fracSolarWindow * transGlass * areaHouse * radSolar[k]) / 1000  # [kWh]
    qRoof[k] = (Tsolair2[k] - Troom[k]) / (rRoof) * areaHouse / 1000  # [kWh]
    qFloor[k] = (Tground - Troom[k]) / (rFloor) * areaHouse / 1000  # [kWh]
    qSideWall[k] = (Tsolair2[k] - Troom[k]) / (rSideWall) * surfaceHouse / 1000  # [kWh]
    qVent[k] = ACH * (Tsolair2[k] - Troom[k]) * mHouse * cAir / 3600

        # 온실에 유입, 유출되는 열 에너지 총합 양수가 나오면 열 획득, 음수가 나오면 열 손실
        # qTotalHouse[k] = qRoof[k] + qFloor[k] + qSideWall[k] + qVent[k] + qFrontBack[k] + qRad[k] # kWh
    qTotalHouse[k] = qRoof[k] + qFloor[k] + qSideWall[k] + qVent[k] + qRad[k] +qHeating[k] -qCooling[k]   # kWh

    # 난방/냉방 부하 계산
    if Troom[k] < Tset_heating:
        # 난방 필요
        qHeating[k] = ((Tset_heating - Troom[k]) * mHouse * cAir / 3600) + abs(qTotalHouse[k]) if qTotalHouse[k] < 0 else 0
        qCooling[k] = 0
        mode[k] = 1  # 난방 모드
    elif Troom[k] > Tset_cooling:
        # 냉방 필요
        qCooling[k] = ((Troom[k] - Tset_cooling) * mHouse * cAir / 3600) + abs(qTotalHouse[k]) if qTotalHouse[k] > 0 else 0
        qHeating[k] = 0
        mode[k] = 2  # 냉방 모드
    else:
        # 난방/냉방 불필요
        qHeating[k] = 0
        qCooling[k] = 0
        mode[k] = 0  # 중립 모드


# 시간 배열 생성
time = range(len(qRad))

# 1. Solar Radiation (qRad)
plt.figure(figsize=(15, 5))
plt.plot(time, qRad, 'r-', label='Solar Radiation', linewidth=1)
plt.xlabel('Time (hours)')
plt.ylabel('Heat Transfer Rate (kW)')
plt.title('Annual Solar Radiation Heat Gain (qRad)')
plt.grid(True, linestyle='--', alpha=0.7)
plt.legend()
plt.xticks(np.arange(0, 8760+1, 1000))
plt.tight_layout()
plt.show()

# 2. Roof Heat Transfer (qRoof)
plt.figure(figsize=(15, 5))
plt.plot(time, qRoof, 'b-', label='Roof Heat Transfer', linewidth=1)
plt.xlabel('Time (hours)')
plt.ylabel('Heat Transfer Rate (kW)')
plt.title('Annual Roof Heat Transfer (qRoof)')
plt.grid(True, linestyle='--', alpha=0.7)
plt.legend()
plt.xticks(np.arange(0, 8760+1, 1000))
plt.tight_layout()
plt.show()

# 3. Floor Heat Transfer (qFloor)
plt.figure(figsize=(15, 5))
plt.plot(time, qFloor, 'g-', label='Floor Heat Transfer', linewidth=1)
plt.xlabel('Time (hours)')
plt.ylabel('Heat Transfer Rate (kW)')
plt.title('Annual Floor Heat Transfer (qFloor)')
plt.grid(True, linestyle='--', alpha=0.7)
plt.legend()
plt.xticks(np.arange(0, 8760+1, 1000))
plt.tight_layout()
plt.show()

# 4. Side Wall Heat Transfer (qSideWall)
plt.figure(figsize=(15, 5))
plt.plot(time, qSideWall, 'purple', label='Side Wall Heat Transfer', linewidth=1)
plt.xlabel('Time (hours)')
plt.ylabel('Heat Transfer Rate (kW)')
plt.title('Annual Side Wall Heat Transfer (qSideWall)')
plt.grid(True, linestyle='--', alpha=0.7)
plt.legend()
plt.xticks(np.arange(0, 8760+1, 1000))
plt.tight_layout()
plt.show()

# 5. Ventilation Heat Transfer (qVent)
plt.figure(figsize=(15, 5))
plt.plot(time, qVent, 'orange', label='Ventilation Heat Transfer', linewidth=1)
plt.xlabel('Time (hours)')
plt.ylabel('Heat Transfer Rate (kW)')
plt.title('Annual Ventilation Heat Transfer (qVent)')
plt.grid(True, linestyle='--', alpha=0.7)
plt.legend()
plt.xticks(np.arange(0, 8760+1, 1000))
plt.tight_layout()
plt.show()

# 각 성분의 연간 총량 및 통계 출력
components = {
    'Solar Radiation (qRad)': qRad,
    'Roof Heat Transfer (qRoof)': qRoof,
    'Floor Heat Transfer (qFloor)': qFloor,
    'Side Wall Heat Transfer (qSideWall)': qSideWall,
    'Ventilation Heat Transfer (qVent)': qVent
}

print("=== 열전달 성분 분석 ===")
for name, data in components.items():
    print(f"\n{name}")
    print(f"연간 총량: {sum(data):.2f} kWh")
    print(f"평균: {np.mean(data):.2f} kW")
    print(f"최대값: {np.max(data):.2f} kW")
    print(f"최소값: {np.min(data):.2f} kW")