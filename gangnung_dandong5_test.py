from __future__ import print_function
import openpyxl
import math
import pandas as pd
import numpy as np
from matplotlib import pyplot as plt
from CoolProp import AbstractState
from CoolProp.CoolProp import PhaseSI, PropsSI, get_global_param_string
import CoolProp.CoolProp as CoolProp

# 기본 상수 정의
P_atm = 101325  # [K]
T_mean = 300  # [K]
rhoAir = PropsSI("D", "T", T_mean, "P", P_atm, "air")  # 1.1769955883877592 [kg/m3]
rhoWater = PropsSI("D", "T", T_mean, "P", P_atm, "water")  # 996.5569352652021 [kg/m3]
cAir = PropsSI("C", "T", T_mean, "P", P_atm, "air") / 1000  # 1.0063739076641027 [kJ/kgK]
cWater = PropsSI("C", "T", T_mean, "P", P_atm, "water") / 1000  # 4.1806357765560715 [kJ/kgK]
pi = 3.14

# 먼저 시간 배열 정의
hour = np.ones(8760)

# 그 다음 나머지 배열들 초기화
Abs = np.ones(1470)
wavelength = np.linspace(0.3, 15, 1470)
Toutdoor = 300 * np.ones(len(hour))  # [K]
Troom = 300 * np.ones(len(hour))
Vwind = np.ones(8760)

hconv = np.ones(8760)
qTotalHouse = np.zeros(len(hour))  # [W]
qInHouse = np.zeros(len(hour))  # [W]
radSolar = np.zeros(len(hour))
qRad = np.zeros(len(hour))
qRoof = np.zeros(len(hour))  # [W]
qFloor = np.zeros(len(hour))  # [W]
qVent = np.zeros(len(hour))  # [W]
qSideWall = np.zeros(len(hour))  # [W]
qFrontBack = np.zeros(len(hour))
qWindow = np.zeros(len(hour))  # 창문 열손실 [W]
qHeating = np.zeros(len(hour))  # [W]
qCooling = np.zeros(len(hour))  # [W]
PW = np.zeros(len(hour))  # [W]
mode = np.zeros(len(hour))  # [W]

# 엑셀 파일 열기 (openpyxl 사용)
wb = openpyxl.load_workbook('TMY3_Gangnung.xlsx')
sheet = wb.active

# 데이터를 엑셀 파일에서 읽어오기
for i in range(0, len(hour)):
    Toutdoor[i] = sheet.cell(row=i + 1, column=2).value  # 배열은 0부터, 엑셀은 1부터 시작
    Vwind[i] = sheet.cell(row=i + 1, column=3).value
    PW[i] = sheet.cell(row=i + 1, column=4).value
    radSolar[i] = sheet.cell(row=i + 1, column=5).value

# 온실 크기 및 기타 변수 설정
lengthHouse = 70  # [m]
widthHouse = 8.6  # [m]
heightHouse = 4.6  # [m]
fracSolarWindow = 0.85
fracLossWindow = 1
nFloor = 1
transGlass = 0.85
rRoof = 2.0
areaRoof = 696.36
rFloor = 2.0
rSideWall = 2.0
areaFrontBack = 64.5
areaSideWall = 604.8
rGround = 3.0
areaHouse = lengthHouse * widthHouse
surfaceHouse = (widthHouse * heightHouse + lengthHouse * heightHouse) * 2
volumeHouse = areaHouse * heightHouse * nFloor
mHouse = areaHouse * heightHouse * nFloor * rhoAir
ht = 5.7 # 열 관류율 [㎉/㎡·h·℃]
hs = 0.244 # 지표면 전열 계수 [㎉/㎡·h·℃]
fr = 0.3 # 보온 피복재의 열 절감율
hv = 0.2
Agh = 1365.66

# 환기 관련 변수 설정
ACH = 0.5  # Air Changes per Hour (시간당 환기횟수)
ventilation_rate = (ACH * volumeHouse) / 3600  # [m³/s]
m_vent = ventilation_rate * rhoAir  # 환기 공기 질량유량 [kg/s]

# 실내 온도 초기화
Troom_initial = 7 + 273
Troom[0] = Troom_initial
Tground = 10 + 273
Tset_heating = 15 + 273  # 난방 설정 온도 (16°C)
Tset_cooling = 20 + 273  # 냉방 설정 온도 (28°C)

# Tsolair2 계산을 위한 변수 정의
alpha_roof = 0.2  # 지붕 표면 일사흡수율 [-]

def calculate_Tsolair2(k, Toutdoor, radSolar):
    Tsolair2 = Toutdoor[k] + (alpha_roof * radSolar[k]) / 17 # 등가 외기온도 [K]
    return Tsolair2

# Tsolair2 계산
Tsolair2 = np.zeros(len(hour))
for k in range(len(hour)):
    Tsolair2[k] = calculate_Tsolair2(k, Toutdoor, radSolar)

# 열 부하 계산
for k in range(len(hour)):
    # 1. 먼저 현재 시간의 실내온도 계산 (k > 0일 때)
    if k > 0:
        dT = (qTotalHouse[k-1] * 0.001163) / (mHouse * cAir)  # 온도 변화
        Troom[k] = Troom[k-1] + dT  # 새로운 실내온도

    # 2. 계산된 실내온도로 열부하 계산
    qVent[k] = (Tsolair2[k] - Troom[k]) * hv * Agh
    qRad[k] = fracSolarWindow * transGlass * areaHouse * radSolar[k]
    qRoof[k] = (Tsolair2[k] - Troom[k]) * ht * areaRoof * (1 - fr)
    qFloor[k] = (Tground - Troom[k]) * hs * areaHouse
    qSideWall[k] = (Tsolair2[k] - Troom[k]) * ht * areaSideWall * (1 - fr)
    qFrontBack[k] = (Tsolair2[k] - Troom[k]) * ht * areaFrontBack * (1 - fr)

    # 온실에 유입, 유출되는 열 에너지 총합
    qTotalHouse[k] = qRoof[k] + qFloor[k] + qSideWall[k] + qVent[k] + qFrontBack[k] + qRad[k] # kcal/h

    # 난방/냉방 부하 계산
    if Troom[k] < Tset_heating:
        # 난방 필요
        qHeating[k] = ((Tset_heating - Troom[k]) * (mHouse * cAir * 1000)) # W
        qCooling[k] = 0
        mode[k] = 1  # 난방 모드
    elif Troom[k] > Tset_cooling:
        # 냉방 필요
        qCooling[k] = ((Troom[k] - Tset_cooling) * (mHouse * cAir * 1000))  # W
        qHeating[k] = 0
        mode[k] = 2  # 냉방 모드
    else:
        # 난방/냉방 불필요
        qHeating[k] = 0
        qCooling[k] = 0
        mode[k] = 0  # 중립 모드

# 데이터를 월별로 분리
def get_monthly_averages(data, month):
    # 각 월의 시작 인덱스 계산
    start_idx = (month - 1) * 24 * 31  # 한 달을 31일로 가정
    end_idx = start_idx + (24 * 31)

    # 해당 월의 데이터 추출
    monthly_data = data[start_idx:end_idx]

    # 24시간 평균 계산
    hourly_averages = np.zeros(24)
    for hour in range(24):
        # 같은 시간대의 모든 날짜 데이터 평균
        hour_data = monthly_data[hour::24]
        hourly_averages[hour] = np.mean(hour_data)

    return hourly_averages

# 1월, 4월, 7월, 10월의 시간별 평균 부하 계산
jan_loads = get_monthly_averages(qTotalHouse, 1)
apr_loads = get_monthly_averages(qTotalHouse, 4)
jul_loads = get_monthly_averages(qTotalHouse, 7)
oct_loads = get_monthly_averages(qTotalHouse, 10)

# 그래프 그리기
plt.figure(figsize=(12, 8))

# 24시간 데이터 플로팅
hours = range(24)
plt.plot(hours, jan_loads / areaHouse, 'r-s', label='Jan.', markersize=8)
plt.plot(hours, apr_loads / areaHouse, 'y-d', label='Apr.', markersize=8)
plt.plot(hours, jul_loads / areaHouse, 'g-^', label='Jul.', markersize=8)
plt.plot(hours, oct_loads / areaHouse, 'b-o', label='Oct.', markersize=8)

# 그래프 설정
plt.xlabel('Time (hr)', fontsize=12)
plt.ylabel('Heating load (W/m²)', fontsize=12)
plt.grid(True, linestyle='--', alpha=0.7)
plt.legend(fontsize=12)

# x축 설정
plt.xlim(0, 24)
plt.xticks(range(0, 25, 2))

# y축 설정 - 논문 그래프와 비슷한 범위로 설정
plt.ylim(-800, 400)

# 레이아웃 조정
plt.tight_layout()
plt.show()

# 기존 코드 맨 아래에 추가
# 총 열부하 그래프 그리기
plt.figure(figsize=(15, 6))

# 시간 배열 생성
time = range(len(hour))

# 열부하 데이터 플로팅 (단위: kW)
plt.plot(time, qTotalHouse / 1000, 'b-', label='Total Heat Load', linewidth=1)

# 그래프 설정
plt.xlabel('Time (hours)', fontsize=12)
plt.ylabel('Heat Load (kW)', fontsize=12)
plt.title('Annual Total Heat Load Variation', fontsize=14)
plt.grid(True, linestyle='--', alpha=0.7)
plt.legend(fontsize=12)

# x축 설정 (1000시간 단위로 눈금 표시)
plt.xticks(np.arange(0, 8760+1, 1000))

# 레이아웃 조정
plt.tight_layout()
plt.show()

# 난방부하와 냉방부하 그래프 그리기
plt.figure(figsize=(15, 6))

# 시간 배열 생성
time = range(len(hour))

# 열부하 데이터 플로팅 (단위: kW)
plt.plot(time, qTotalHouse / 1000, 'gray', label='Total Heat Load', alpha=0.5, linewidth=1)
plt.plot(time, qHeating / 1000, 'r-', label='Heating Load', linewidth=1)
plt.plot(time, qCooling / 1000, 'b-', label='Cooling Load', linewidth=1)

# 그래프 설정
plt.xlabel('Time (hours)', fontsize=12)
plt.ylabel('Heat Load (kW)', fontsize=12)
plt.title('Annual Heating and Cooling Load Variation', fontsize=14)
plt.grid(True, linestyle='--', alpha=0.7)
plt.legend(fontsize=12)

# x축 설정
plt.xticks(np.arange(0, 8760+1, 1000))

# 레이아웃 조정
plt.tight_layout()
plt.show()

# Add new visualization for room temperature
plt.figure(figsize=(15, 6))

# Create time array
time = range(len(hour))

# Plot room temperature (convert from Kelvin to Celsius)
plt.plot(time, Troom - 273.15, 'g-', label='Room Temperature', linewidth=1)

# Add horizontal lines for heating and cooling setpoints
plt.axhline(y=Tset_heating - 273.15, color='r', linestyle='--', label='Heating Setpoint')
plt.axhline(y=Tset_cooling - 273.15, color='b', linestyle='--', label='Cooling Setpoint')

# Graph settings
plt.xlabel('Time (hours)', fontsize=12)
plt.ylabel('Temperature (°C)', fontsize=12)
plt.title('Annual Room Temperature Variation', fontsize=14)
plt.grid(True, linestyle='--', alpha=0.7)
plt.legend(fontsize=12)

# X-axis settings (tick marks every 1000 hours)
plt.xticks(np.arange(0, 8760+1, 1000))

# Adjust layout
plt.tight_layout()
plt.show()

# 난방부하와 냉방부하를 별도의 그래프로 표시
fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(15, 10))

# 난방부하 그래프
ax1.plot(time, qHeating / 1000, 'r-', label='Heating Load', linewidth=1)
ax1.set_xlabel('Time (hours)', fontsize=12)
ax1.set_ylabel('Heating Load (kW)', fontsize=12)
ax1.set_title('Annual Heating Load Variation', fontsize=14)
ax1.grid(True, linestyle='--', alpha=0.7)
ax1.legend(fontsize=12)
ax1.set_xticks(np.arange(0, 8760+1, 1000))

# 냉방부하 그래프
ax2.plot(time, qCooling / 1000, 'b-', label='Cooling Load', linewidth=1)
ax2.set_xlabel('Time (hours)', fontsize=12)
ax2.set_ylabel('Cooling Load (kW)', fontsize=12)
ax2.set_title('Annual Cooling Load Variation', fontsize=14)
ax2.grid(True, linestyle='--', alpha=0.7)
ax2.legend(fontsize=12)
ax2.set_xticks(np.arange(0, 8760+1, 1000))

plt.tight_layout()
plt.show()

# Tsolair2, Troom, Toutdoor 온도 비교 그래프 생성
plt.figure(figsize=(15, 6))

# 시간 배열 생성
time = range(len(hour))

# 온도 데이터 플로팅 (켈빈에서 섭씨로 변환)
plt.plot(time, Tsolair2 - 273.15, 'r-', label='Equivalent Outdoor Temperature (Tsolair2)', linewidth=1)
plt.plot(time, Troom - 273.15, 'g-', label='Room Temperature', linewidth=1)

# 그래프 설정
plt.xlabel('Time (hours)', fontsize=12)
plt.ylabel('Temperature (°C)', fontsize=12)
plt.title('Annual Temperature Variation Comparison', fontsize=14)
plt.grid(True, linestyle='--', alpha=0.7)
plt.legend(fontsize=10, loc='best')

# x축 설정 (1000시간 단위로 눈금 표시)
plt.xticks(np.arange(0, 8760+1, 1000))

# y축 범위 설정
plt.ylim(min(min(Tsolair2 - 273.15), min(Toutdoor - 273.15)) - 5,
         max(max(Tsolair2 - 273.15), max(Toutdoor - 273.15)) + 5)

# 레이아웃 조정
plt.tight_layout()
plt.show()

# 연간 부하 출력
print(f"Annual Total Heat Load: {sum(abs(qTotalHouse)) / 1000:.2f} kWh")
print(f"Annual Heating Load: {sum(qHeating) / 1000:.2f} kWh")
print(f"Annual Cooling Load: {sum(qCooling) / 1000:.2f} kWh")

# 최대 난방부하 계산 및 출력
max_heating_load = max(qHeating) / 1000   # kW 단위로 변환
max_heating_hour = np.argmax(qHeating) + 1  # 최대 난방부하가 발생한 시간
print(f"\nMaximum Heating Load: {max_heating_load:.2f} kW (Hour {max_heating_hour})")

# 최저 외기온도 계산 및 출력
min_solair2_temp_K = min(Tsolair2)
min_solair2_temp_C = min_solair2_temp_K - 273.15  # 켈빈 온도를 섭씨로 변환
min_temp_hour = np.argmin(Tsolair2) + 1  # 최저 기온이 발생한 시간
print(f"Minimum Outdoor Temperature: {min_solair2_temp_C:.2f}°C (Hour {min_temp_hour})")

Troom_C = Troom - 273.15  # 켈빈을 섭씨로 변환
print("\n=== 실내온도 ===")
print(f"평균: {np.mean(Troom_C):.2f}°C")
print(f"최대: {np.max(Troom_C):.2f}°C")
print(f"최소: {np.min(Troom_C):.2f}°C")

# 최대 난방부하가 발생한 시점의 조건 확인
max_heating_idx = np.argmax(qHeating)
print(f"\n=== 최대 난방부하 발생 시점 분석 ===")
print(f"시간: {max_heating_idx}")
print(f"실내온도: {Troom[max_heating_idx]-273.15:.2f}°C")
print(f"외기온도: {Tsolair2[max_heating_idx]-273.15:.2f}°C")
print(f"난방부하: {qHeating[max_heating_idx] / 1000:.2f}kW")

# 최저 실내온도 발생 시점의 조건 확인
min_temp_idx = np.argmin(Troom)
print(f"\n=== 최저 실내온도 발생 시점 분석 ===")
print(f"시간: {min_temp_idx}")
print(f"실내온도: {Troom[min_temp_idx]-273.15:.2f}°C")
print(f"외기온도: {Tsolair2[min_temp_idx]-273.15:.2f}°C")
print(f"난방부하: {qHeating[min_temp_idx] / 1000:.2f}kW")

# for k in range(len(hour)):
#     print(Troom[k])