from __future__ import print_function
import openpyxl
import math
import pandas as pd
import numpy as np
from matplotlib import pyplot as plt
from CoolProp import AbstractState
from CoolProp.CoolProp import PhaseSI, PropsSI, get_global_param_string
import CoolProp.CoolProp as CoolProp

# 기본 상수 정의
P_atm = 101325  # [K]
T_mean = 300  # [K]
rhoAir = PropsSI("D", "T", T_mean, "P", P_atm, "air")  # [kg/m3]
rhoWater = PropsSI("D", "T", T_mean, "P", P_atm, "water")  # [kg/m3]
cAir = PropsSI("C", "T", T_mean, "P", P_atm, "air") / 1000  # [kJ/kgK]
cWater = PropsSI("C", "T", T_mean, "P", P_atm, "water") / 1000  # [kJ/kgK]
pi = 3.14

# 먼저 시간 배열 정의
hour = np.ones(8760)

# 그 다음 나머지 배열들 초기화
Abs = np.ones(1470)
wavelength = np.linspace(0.3, 15, 1470)
Toutdoor = 300 * np.ones(len(hour))  # [K]
Troom = 295
# Troom = 300 * np.ones(len(hour))
Vwind = np.ones(8760)
hconv = np.ones(8760)
qTotalHouse = np.zeros(len(hour))  # [W]
qInHouse = np.zeros(len(hour))  # [W]
radSolar = np.zeros(len(hour))
qRad = np.zeros(len(hour))
qRoof = np.zeros(len(hour))  # [W]
qFloor = np.zeros(len(hour))  # [W]
qVent = np.zeros(len(hour))  # [W]
qSideWall = np.zeros(len(hour))  # [W]
qFrontBack = np.zeros(len(hour))
qWindow = np.zeros(len(hour))  # 창문 열손실 [W]
qHeating = np.zeros(len(hour))  # [W]
qCooling = np.zeros(len(hour))  # [W]
PW = np.zeros(len(hour))  # [W]
mode = np.zeros(len(hour))  # [W]

# Tsolair2 계산을 위한 변수 정의
alpha_roof = 0.2  # 지붕 표면 일사흡수율 [-]
epsilon = 0.9  # 지붕 표면 방사율 [-]
sigma = 5.67e-8  # Stefan-Boltzmann 상수 [W/m2K4]

# Tsolair2 계산 함수
def calculate_Tsolair2(k, Toutdoor, radSolar, Vwind):
    """
    등가 외기온도(sol-air temperature) 계산
    """
    h_out_corrected = 5.7 + 3.8 * Vwind[k]  # 풍속 보정된 열전달계수 [W/m2K]
    dT_longwave = (epsilon * 63) / h_out_corrected  # 장파장 복사 보정항 [K]
    Tsolair2 = Toutdoor[k] + (alpha_roof * radSolar[k]) / h_out_corrected - dT_longwave  # 등가 외기온도 [K]
    return Tsolair2

# 엑셀 파일 열기 (openpyxl 사용)
wb = openpyxl.load_workbook('TMY3_Gangnung.xlsx')
sheet = wb.active

# 데이터를 엑셀 파일에서 읽어오기
for i in range(0, len(hour)):
    Toutdoor[i] = sheet.cell(i + 1, 2).value
    Vwind[i] = sheet.cell(i + 1, 3).value
    PW[i] = sheet.cell(i + 1, 4).value
    radSolar[i] = sheet.cell(i + 1, 5).value

# 온실 크기 및 기타 변수 설정
lengthHouse = 70  # [m]
widthHouse = 8.6  # [m]
heightHouse = 4.6  # [m]
fracSolarWindow = 0.5
fracLossWindow = 1
nFloor = 1
transGlass = 0.8
rRoof = 2.0
areaRoof = 2217
rFloor = 2.0
rSideWall = 2.0
areaFrontBack = 300
areaSideWall = 385
rGround = 3.0
areaHouse = lengthHouse * widthHouse
surfaceHouse = (widthHouse * heightHouse + lengthHouse * heightHouse) * 2
volumeHouse = areaHouse * heightHouse * nFloor
mHouse = areaHouse * heightHouse * nFloor * rhoAir
ht = 5.7 # 열 관류율 [㎉/㎡·h·℃]
hs = 0.244 # 지표면 전열 계수 [㎉/㎡·h·℃]
areasoil = 2200 # 온실 면적 [㎡]
fr = 0.3 # 보온 피복재의 열 절감율
hv = 0.2
Agh = 2902
# 환기 관련 변수 설정
ACH = 0.5  # Air Changes per Hour (시간당 환기횟수)
ventilation_rate = (ACH * volumeHouse) / 3600  # [m³/s]
m_vent = ventilation_rate * rhoAir  # 환기 공기 질량유량 [kg/s]

# 실내 온도 초기화
Troom_initial = 295
Troom = Troom_initial
Tground = 4.12 + 273
Tset_heating = 16 + 273  # 난방 설정 온도 (20°C)
Tset_cooling = 28 + 273  # 냉방 설정 온도 (28°C)

# Tsolair2 계산
Tsolair2 = np.zeros(len(hour))
for k in range(len(hour)):
    Tsolair2[k] = calculate_Tsolair2(k, Toutdoor, radSolar, Vwind)

# 열 부하 계산
for k in range(len(hour)):
    # 환기 전열부하 계산 [W]
    # qVent[k] = m_vent * cAir * 1000 * (Troom - Toutdoor[k])  # cAir를 kJ/kgK에서 J/kgK로 변환
    qVent[k] = (Troom - Toutdoor[k]) * hv * Agh
    qRad[k] = fracSolarWindow * transGlass * areaHouse * radSolar[k]
    qRoof[k] = (Troom - Toutdoor[k]) * ht * areaRoof * (1 - fr)
    qFloor[k] = (Troom - Tground) * hs * areasoil
    qSideWall[k] = (Troom - Toutdoor[k]) * ht * areaSideWall * (1 - fr)
    qFrontBack[k] = (Troom - Toutdoor[k]) * ht * areaFrontBack * (1 - fr)

    # 총 열부하 계산 (환기부하 포함)
    qTotalHouse[k] = qRoof[k] + qFloor[k] + qSideWall[k] + qVent[k] + qFrontBack[k] - qRad[k]

    # 난방/냉방 부하 계산
    if Toutdoor[k] < Tset_heating:
        qHeating[k] = abs(qTotalHouse[k])  # 난방 필요
        qCooling[k] = 0
    elif Toutdoor[k] > Tset_cooling:
        qCooling[k] = abs(qTotalHouse[k])  # 냉방 필요
        qHeating[k] = 0
    else:
        qHeating[k] = 0
        qCooling[k] = 0

# 기존 코드는 그대로 유지하고 그래프 부분만 수정
# 데이터를 월별로 분리
def get_monthly_averages(data, month):
    # 각 월의 시작 인덱스 계산
    start_idx = (month - 1) * 24 * 31  # 한 달을 31일로 가정
    end_idx = start_idx + (24 * 31)

    # 해당 월의 데이터 추출
    monthly_data = data[start_idx:end_idx]

    # 24시간 평균 계산
    hourly_averages = np.zeros(24)
    for hour in range(24):
        # 같은 시간대의 모든 날짜 데이터 평균
        hour_data = monthly_data[hour::24]
        hourly_averages[hour] = np.mean(hour_data)

    return hourly_averages


# 1월, 4월, 7월, 10월의 시간별 평균 부하 계산
jan_loads = get_monthly_averages(qTotalHouse, 1)
apr_loads = get_monthly_averages(qTotalHouse, 4)
jul_loads = get_monthly_averages(qTotalHouse, 7)
oct_loads = get_monthly_averages(qTotalHouse, 10)



# 그래프 그리기
plt.figure(figsize=(12, 8))

# 24시간 데이터 플로팅
hours = range(24)
plt.plot(hours, jan_loads / areaHouse, 'r-s', label='Jan.', markersize=8)
plt.plot(hours, apr_loads / areaHouse, 'y-d', label='Apr.', markersize=8)
plt.plot(hours, jul_loads / areaHouse, 'g-^', label='Jul.', markersize=8)
plt.plot(hours, oct_loads / areaHouse, 'b-o', label='Oct.', markersize=8)



# 그래프 설정
plt.xlabel('Time (hr)', fontsize=12)
plt.ylabel('Heating load (W/m²)', fontsize=12)
plt.grid(True, linestyle='--', alpha=0.7)
plt.legend(fontsize=12)

# x축 설정
plt.xlim(0, 24)
plt.xticks(range(0, 25, 2))

# y축 설정 - 논문 그래프와 비슷한 범위로 설정
plt.ylim(-400, 800)

# Heating/Cooling required 텍스트 추가
plt.text(2, 300, 'Heating required', fontsize=12)
plt.text(2, -300, 'Cooling required', fontsize=12)

# 레이아웃 조정
plt.tight_layout()
plt.show()

# 연간 부하 출력
print(f"Annual Total Heat Load: {sum(abs(qTotalHouse)) / (1000):.2f} kWh")
print(f"Annual Heating Load: {sum(qHeating) / (1000):.2f} kWh")
print(f"Annual Cooling Load: {sum(qCooling) / (1000):.2f} kWh")
print(f"Annual Ventilation Load: {sum(abs(qVent)) / (1000):.2f} kWh")

# 최대 난방부하 계산 및 출력
max_heating_load = max(qHeating) / 1000  # kW 단위로 변환
max_heating_hour = np.argmax(qHeating) + 1  # 최대 난방부하가 발생한 시간
print(f"\nMaximum Heating Load: {max_heating_load:.2f} kW (Hour {max_heating_hour})")

# 최저 외기온도 계산 및 출력
min_outdoor_temp_K = min(Toutdoor)
min_outdoor_temp_C = min_outdoor_temp_K - 273.15  # 켈빨 온도를 섭씨로 변환
min_temp_hour = np.argmin(Toutdoor) + 1  # 최저 기온이 발생한 시간
print(f"Minimum Outdoor Temperature: {min_outdoor_temp_C:.2f}°C (Hour {min_temp_hour})")
