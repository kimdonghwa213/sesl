## 김제 스마트팜 혁신밸리 난방부하 계산

from __future__ import print_function
import xlrd
import openpyxl
import math
import pandas as pd
import numpy as np
from matplotlib import pyplot as plt
from CoolProp import AbstractState
from CoolProp.CoolProp import PhaseSI, PropsSI, get_global_param_string
import CoolProp.CoolProp as CoolProp

# 기본 물성치 정의
P_atm = 101325  # 대기압 [Pa]
T_mean = 300  # 평균 온도 [K]
rhoAir = PropsSI("D", "T", T_mean, "P", P_atm, "air")  # 공기 밀도 [kg/m3]
rhoWater = PropsSI("D", "T", T_mean, "P", P_atm, "water")  # 물 밀도 [kg/m3]
cAir = PropsSI("C", "T", T_mean, "P", P_atm, "air") / 1000  # 공기 비열 [kJ/kgK]
cWater = PropsSI("C", "T", T_mean, "P", P_atm, "water") / 1000  # 물 비열 [kJ/kgK]

# 지붕 재료 물성치 정의 (유리로 변경)
k_glass = 0.96  # 유리 열전도율 [W/mK]
t_glass_roof = 0.012  # 지붕 유리 두께 [m]
h_in = 8.3  # 실내측 열전달계수 [W/m2K]
h_out = 23.0  # 실외측 열전달계수 [W/m2K]

# 바닥 재료 물성치 정의
k_soil = 1.5  # 토양 열전도율 [W/mK]
t_soil = 0.5  # 토양 두께 [m]
k_slab = 1.13  # 콘크리트 슬래브 열전도율 [W/mK]
t_slab = 0.15  # 슬래브 두께 [m]

# 벽체 재료 물성치 정의
t_glass_wall = 0.010  # 벽체 유리 두께 [m]

# 각 부위별 열저항 계산 [m2K/W]
rRoof = (1 / h_in) + (t_glass_roof / k_glass) + (1 / h_out)  # 유리 지붕 총 열저항
rFloor = (1 / h_in) + (t_slab / k_slab) + (t_soil / k_soil)  # 바닥 총 열저항
rSideWall = (1 / h_in) + (t_glass_wall / k_glass) + (1 / h_out)  # 유리 벽체 총 열저항

# Tsolair2 계산을 위한 변수 정의
alpha_roof = 0.2  # 지붕 표면 일사흡수율 [-]
epsilon = 0.9  # 지붕 표면 방사율 [-]
sigma = 5.67e-8  # Stefan-Boltzmann 상수 [W/m2K4]

pi = 3.14
hour = np.ones(8760)  # 연간 시간 배열 [hr]
Abs = np.ones(1470)  # 흡수율 배열 [-]
wavelength = np.linspace(0.3, 15, 1470)  # 파장 범위 [μm]
Toutdoor = 300 * np.ones(len(hour))  # 외기온도 배열 [K]
Troom = 300 * np.ones(len(hour))  # 실내온도 배열 [K]
Vwind = np.ones(8760)  # 풍속 배열 [m/s]

# 열전달 관련 배열 초기화
hconv = np.ones(8760)  # 대류 열전달계수 [W/m2K]
qTotalHouse = np.zeros(len(hour))  # 총 열부하 [W]
qInHouse = np.zeros(len(hour))  # 실내 발열 [W]
radSolar = np.zeros(len(hour))  # 일사량 [W/m2]
qRad = np.zeros(len(hour))  # 복사 열전달량 [W]
qRoof = np.zeros(len(hour))  # 지붕 열전달량 [W]
qFloor = np.zeros(len(hour))  # 바닥 열전달량 [W]
qVent = np.zeros(len(hour))  # 환기 열전달량 [W]
qSideWall = np.zeros(len(hour))  # 벽체 열전달량 [W]
qHeating = np.zeros(len(hour))  # 난방 부하 [W]
qCooling = np.zeros(len(hour))  # 냉방 부하 [W]
PW = np.zeros(len(hour))  # 수증기분압 [Pa]
mode = np.zeros(len(hour))  # 운전 모드 (0: 정지, 1: 난방, 2: 냉방)

# Tsolair2 계산 함수
def calculate_Tsolair2(k, Toutdoor, radSolar, Vwind):
    """
    등가 외기온도(sol-air temperature) 계산
    """
    h_out_corrected = 5.7 + 3.8 * Vwind[k]  # 풍속 보정된 열전달계수 [W/m2K]
    dT_longwave = (epsilon * 63) / h_out_corrected  # 장파장 복사 보정항 [K]
    Tsolair2 = Toutdoor[k] + (alpha_roof * radSolar[k]) / h_out_corrected - dT_longwave  # 등가 외기온도 [K]
    return Tsolair2

#엑셀 파일 열기 (openpyxl 사용)
wb = openpyxl.load_workbook('TMY3_Gangnung.xlsx')
sheet = wb.active

# 데이터를 엑셀 파일에서 읽어오기
for i in range(0, len(hour)):
    Toutdoor[i] = sheet.cell(i + 1, 2).value
    Vwind[i] = sheet.cell(i + 1, 3).value
    PW[i] = sheet.cell(i + 1, 4).value
    radSolar[i] = sheet.cell(i + 1, 5).value

# 건물 치수 정의
lengthHouse = 70  # 건물 길이 [m]
widthHouse = 8.6  # 건물 폭 [m]
heightHouse = 4.6  # 건물 높이 [m]
fracSolarWindow = 0.5  # 창문 일사 투과율 [-]
fracLossWindow = 1  # 창문 열손실 계수 [-]
nFloor = 1  # 층수 [-]
areaHouse = lengthHouse * widthHouse  # 건물 바닥면적 [m2]
surfaceHouse = (widthHouse * heightHouse + lengthHouse * heightHouse) * 2  # 건물 외피면적 [m2]
volumeHouse = areaHouse * heightHouse * nFloor  # 건물 체적 [m3]
mHouse = areaHouse * heightHouse * nFloor * rhoAir  # 건물 내 공기질량 [kg]

# 실내 환경 설정
Troom_initial = 297  # 초기 실내온도 [K]
Troom[0] = Troom_initial  # 실내온도 초기화 [K]
Tground = 10 + 273  # 지중온도 [K]
error = 10  # 허용오차 [K]
Tset_heating = 10 + 273  # 난방설정온도 [K]
Tset_cooling = 34 + 273  # 냉방설정온도 [K]

# 창문 특성
transGlass = 0.8  # 유리 투과율 [-]

# 시간 간격 설정
dt = 3600  # 시간 간격 (1시간 = 3600초)
cp_air = 1.005  # 공기의 비열 [kJ/kg·K]

# Tsolair2 계산
Tsolair2 = np.zeros(len(hour))
for k in range(len(hour)):
    Tsolair2[k] = calculate_Tsolair2(k, Toutdoor, radSolar, Vwind)

# 열부하 계산
for k in range(len(hour)):
    qRad[k] = (fracSolarWindow * transGlass * areaHouse * radSolar[k]) / 1000  # 일사 열획득 [kWh]
    qRoof[k] = (Tsolair2[k] - Troom[k]) / (rRoof) * areaHouse / 1000  # 지붕 열전달 [kWh]
    qFloor[k] = (Tground - Troom[k]) / (rFloor) * areaHouse / 1000  # 바닥 열전달 [kWh]
    qSideWall[k] = (Tsolair2[k] - Troom[k]) / (rSideWall) * surfaceHouse / 1000  # 벽체 열전달 [kWh]

# 총 열부하 계산
qTotalHouse = qRad + qRoof + qFloor + qSideWall  # 총 열부하 [kWh]

# 실내온도 변화 계산
for k in range(1, len(hour)):
    # 이전 시간의 열부하로 현재 실내온도 계산
    dT = (qTotalHouse[k-1] * 1000) / (mHouse * cp_air)  # [K]
    Troom[k] = Troom[k-1] + dT

# 난방/냉방 부하 계산
for k in range(len(hour)):
    if Troom[k] < Tset_heating:
        # 난방이 필요한 경우
        qHeating[k] = (Tset_heating - Troom[k]) * mHouse * cp_air / (dt * 1000)  # [kWh]
        qCooling[k] = 0
        mode[k] = 1  # 난방모드
    elif Troom[k] > Tset_cooling:
        # 냉방이 필요한 경우
        qCooling[k] = (Troom[k] - Tset_cooling) * mHouse * cp_air / (dt * 1000)  # [kWh]
        qHeating[k] = 0
        mode[k] = 2  # 냉방모드
    else:
        # 난방/냉방 불필요
        qHeating[k] = 0
        qCooling[k] = 0
        mode[k] = 0  # 정지모드

# 그래프 생성
plt.figure(figsize=(15, 15))

# 1. 실내온도 변화 그래프
plt.subplot(4, 1, 1)
plt.plot(range(len(hour)), [t-273 for t in Troom], 'g-', label='Room Temperature')
plt.axhline(y=Tset_heating-273, color='r', linestyle='--', label='Heating Setpoint')
plt.axhline(y=Tset_cooling-273, color='b', linestyle='--', label='Cooling Setpoint')
plt.title('Room Temperature Profile', fontsize=12, pad=10)
plt.xlabel('Hour of Year')
plt.ylabel('Temperature (°C)')
plt.grid(True, linestyle='--', alpha=0.7)
plt.legend()

# 2. 총 열부하 그래프
plt.subplot(4, 1, 2)
plt.plot(range(len(hour)), qTotalHouse, 'b-', label='Total Heat Load')
plt.title('Total Heat Load Profile', fontsize=12, pad=10)
plt.xlabel('Hour of Year')
plt.ylabel('Heat Load (kWh)')
plt.grid(True, linestyle='--', alpha=0.7)
plt.legend()

# 3. 난방/냉방 부하 그래프
plt.subplot(4, 1, 3)
plt.plot(range(len(hour)), qHeating, 'r-', label='Heating Load')
plt.plot(range(len(hour)), qCooling, 'b-', label='Cooling Load')
plt.title('Heating and Cooling Load Profile', fontsize=12, pad=10)
plt.xlabel('Hour of Year')
plt.ylabel('Load (kWh)')
plt.grid(True, linestyle='--', alpha=0.7)
plt.legend()

# 4. 구성요소별 열부하 그래프
plt.subplot(4, 1, 4)
plt.plot(range(len(hour)), qRoof, 'r-', label='Roof Load', alpha=0.7)
plt.plot(range(len(hour)), qFloor, 'g-', label='Floor Load', alpha=0.7)
plt.plot(range(len(hour)), qSideWall, 'b-', label='Wall Load', alpha=0.7)
plt.plot(range(len(hour)), qRad, 'y-', label='Solar Load', alpha=0.7)
plt.title('Component-wise Heat Load Profile', fontsize=12, pad=10)
plt.xlabel('Hour of Year')
plt.ylabel('Load (kWh)')
plt.grid(True, linestyle='--', alpha=0.7)
plt.legend()

plt.tight_layout()
plt.show()

# 연간 총량 계산 및 출력
print("\nAnnual Heat Load Summary:")
print(f"Total Heat Load: {sum(abs(qTotalHouse)):.2f} kWh")
print(f"Heating Load: {sum(qHeating):.2f} kWh")
print(f"Cooling Load: {sum(qCooling):.2f} kWh")