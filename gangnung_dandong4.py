from __future__ import print_function
import openpyxl
import math
import pandas as pd
import numpy as np
from matplotlib import pyplot as plt
from CoolProp import AbstractState
from CoolProp.CoolProp import PhaseSI, PropsSI, get_global_param_string
import CoolProp.CoolProp as CoolProp

# 기본 상수 정의
P_atm = 101325  # [K]
T_mean = 300  # [K]
rhoAir = PropsSI("D", "T", T_mean, "P", P_atm, "air")  # [kg/m3]
rhoWater = PropsSI("D", "T", T_mean, "P", P_atm, "water")  # [kg/m3]
cAir = PropsSI("C", "T", T_mean, "P", P_atm, "air") / 1000  # [kJ/kgK]
cWater = PropsSI("C", "T", T_mean, "P", P_atm, "water") / 1000  # [kJ/kgK]
pi = 3.14

# 먼저 시간 배열 정의
hour = np.ones(8760)

# 그 다음 나머지 배열들 초기화
Abs = np.ones(1470)
wavelength = np.linspace(0.3, 15, 1470)
Toutdoor = 300 * np.ones(len(hour))  # [K]
Troom = 300 * np.ones(len(hour))
Vwind = np.ones(8760)
hconv = np.ones(8760)
qTotalHouse = np.zeros(len(hour))  # [W]
qInHouse = np.zeros(len(hour))  # [W]
radSolar = np.zeros(len(hour))
qRad = np.zeros(len(hour))
qRoof = np.zeros(len(hour))  # [W]
qFloor = np.zeros(len(hour))  # [W]
qVent = np.zeros(len(hour))  # [W]
qSideWall = np.zeros(len(hour))  # [W]
qWindow = np.zeros(len(hour))  # 창문 열손실 [W]
qHeating = np.zeros(len(hour))  # [W]
qCooling = np.zeros(len(hour))  # [W]
PW = np.zeros(len(hour))  # [W]
mode = np.zeros(len(hour))  # [W]

# 엑셀 파일 열기 (openpyxl 사용)
wb = openpyxl.load_workbook('TMY3_Gangnung.xlsx')
sheet = wb.active

# 데이터를 엑셀 파일에서 읽어오기
for i in range(0, len(hour)):
    Toutdoor[i] = sheet.cell(i + 1, 2).value
    Vwind[i] = sheet.cell(i + 1, 3).value
    PW[i] = sheet.cell(i + 1, 4).value
    radSolar[i] = sheet.cell(i + 1, 5).value

# 온실 크기 및 기타 변수 설정
lengthHouse = 70  # [m]
widthHouse = 8.6  # [m]
heightHouse = 4.6  # [m]
fracSolarWindow = 0.5
fracLossWindow = 1
nFloor = 1
transGlass = 0.8
rRoof = 2.0
rFloor = 2.0
rSideWall = 2.0
rWindow = 0.8
areaWindow = 15
rDoor = 1.5
areaDoor = 5
rGround = 3.0
Tsolair2 = np.ones(8760) * 300
areaHouse = lengthHouse * widthHouse
surfaceHouse = (widthHouse * heightHouse + lengthHouse * heightHouse) * 2
volumeHouse = areaHouse * heightHouse * nFloor
mHouse = areaHouse * heightHouse * nFloor * rhoAir

# 환기 관련 변수 설정
ACH = 0.5  # Air Changes per Hour (시간당 환기횟수)
ventilation_rate = (ACH * volumeHouse) / 3600  # [m³/s]
m_vent = ventilation_rate * rhoAir  # 환기 공기 질량유량 [kg/s]

# 실내 온도 초기화
Troom_initial = 297
Troom[0] = Troom_initial
Tground = 10 + 273
Tset_heating = 20 + 273  # 난방 설정 온도 (20°C)
Tset_cooling = 28 + 273  # 냉방 설정 온도 (28°C)

# 열 부하 계산
for k in range(len(hour)):
    # 환기 전열부하 계산 [W]
    qVent[k] = m_vent * cAir * 1000 * (Toutdoor[k] - Troom[k])  # cAir를 kJ/kgK에서 J/kgK로 변환

    qRad[k] = fracSolarWindow * transGlass * areaHouse * radSolar[k]
    qRoof[k] = (Tsolair2[k] - Troom[k]) / rRoof * areaHouse
    qFloor[k] = (Tground - Troom[k]) / rFloor * areaHouse
    qSideWall[k] = (Tsolair2[k] - Troom[k]) / rSideWall * surfaceHouse
    qWindow[k] = (Toutdoor[k] - Troom[k]) / rWindow * areaWindow
    qDoor = (Toutdoor[k] - Troom[k]) / rDoor * areaDoor

    # 총 열부하 계산 (환기부하 포함)
    qTotalHouse[k] = qRad[k] + qRoof[k] + qFloor[k] + qSideWall[k] + qWindow[k] + qDoor + qVent[k]

    # 난방/냉방 부하 계산
    if Toutdoor[k] < Tset_heating:
        qHeating[k] = abs(qTotalHouse[k])  # 난방 필요
        qCooling[k] = 0
    elif Toutdoor[k] > Tset_cooling:
        qCooling[k] = abs(qTotalHouse[k])  # 냉방 필요
        qHeating[k] = 0
    else:
        qHeating[k] = 0
        qCooling[k] = 0

# 그래프 시각화
plt.figure(figsize=(15, 15))

# 서브플롯 1: 전체 열부하
plt.subplot(3, 1, 1)
plt.plot(range(8760), qTotalHouse / 1000, 'b-', label='Total Heat Load')
plt.title('Greenhouse Total Heat Load Profile')
plt.xlabel('Hour of Year')
plt.ylabel('Heat Load (kW)')
plt.grid(True)
plt.legend()

# 서브플롯 2: 난방 및 냉방 부하
plt.subplot(3, 1, 2)
plt.plot(range(8760), qHeating / 1000, 'r-', label='Heating Load')
plt.plot(range(8760), qCooling / 1000, 'b-', label='Cooling Load')
plt.title('Greenhouse Heating and Cooling Load Profile')
plt.xlabel('Hour of Year')
plt.ylabel('Load (kW)')
plt.grid(True)
plt.legend()

# 서브플롯 3: 환기 전열부하
plt.subplot(3, 1, 3)
plt.plot(range(8760), qVent / 1000, 'g-', label='Ventilation Load')
plt.title('Greenhouse Ventilation Heat Load Profile')
plt.xlabel('Hour of Year')
plt.ylabel('Load (kW)')
plt.grid(True)
plt.legend()

plt.tight_layout()
plt.show()

# 연간 부하 출력
print(f"Annual Total Heat Load: {sum(abs(qTotalHouse)) / (1000):.2f} kWh")
print(f"Annual Heating Load: {sum(qHeating) / (1000):.2f} kWh")
print(f"Annual Cooling Load: {sum(qCooling) / (1000):.2f} kWh")
print(f"Annual Ventilation Load: {sum(abs(qVent)) / (1000):.2f} kWh")